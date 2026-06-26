import os
import re
import json
import time
import urllib.request
import urllib.parse
from flask import Flask, jsonify, request, render_template, send_from_directory
import openpyxl
import google.generativeai as genai

app = Flask(__name__, template_folder='templates', static_folder='static')

# Configurações
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Marmitas da Lulu - Combo 20 unidades.xlsx")
BASE_URL = "https://marmitasdalulu.com.br"
CLIENTE_PATH = "marmitasdalulu"
DEFAULT_COMBO_ID = "45"
REQUEST_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}
COMBOS_CACHE_TTL_SECONDS = 60

FALLBACK_COMBOS = {
    "47": {
        "id": 47,
        "nome": 'MONTE SEU COMBO "TÔ NO FOCO" A PARTIR DE 10 UNIDADES',
        "rotulo": 'Combo "TÔ NO FOCO" - 10 marmitas',
        "slug": "monte-seu-combo--to-no-foco--a-partir-de-10-unidades",
        "min_quantidade": 10,
        "quantidade_padrao": 10,
        "max_quantidade": 1000,
        "preco_a_partir": 267.80,
        "preco_a_partir_texto": "A partir de R$ 267,80",
    },
    "45": {
        "id": 45,
        "nome": 'MONTE SEU COMBO "VIDA SAUDÁVEL" A PARTIR DE 20 UNIDADES',
        "rotulo": 'Combo "VIDA SAUDÁVEL" - 20 marmitas',
        "slug": "monte-seu-combo--vida-saudavel--a-partir-de-20-unidades",
        "min_quantidade": 20,
        "quantidade_padrao": 20,
        "max_quantidade": 100,
        "preco_a_partir": 524.60,
        "preco_a_partir_texto": "A partir de R$ 524,60",
    },
    "544": {
        "id": 544,
        "nome": 'MONTE SEU COMBO "ESTILO DE VIDA" A PARTIR DE 40 UNIDADES',
        "rotulo": 'Combo "ESTILO DE VIDA" - 40 marmitas',
        "slug": "monte-seu-combo--estilo-de-vida--a-partir-de-40-unidades",
        "min_quantidade": 40,
        "quantidade_padrao": 40,
        "max_quantidade": 1000,
        "preco_a_partir": 1026.80,
        "preco_a_partir_texto": "A partir de R$ 1.026,80",
    },
}

FALLBACK_COMBO_ORDER = ["47", "45", "544"]
SITE_DATA_CACHE = {
    "produtos": None,
    "expires_at": 0,
}

for combo in FALLBACK_COMBOS.values():
    combo["url"] = f"{BASE_URL}/{CLIENTE_PATH}/produto/{combo['id']}/{combo['slug']}"


def formatar_preco_brasileiro(valor):
    if valor is None:
        return ""
    valor_formatado = f"{float(valor):,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"R$ {valor_formatado}"


def montar_combo_url(combo_id, slug):
    return f"{BASE_URL}/{CLIENTE_PATH}/produto/{combo_id}/{slug}"


def obter_composicao_principal(combo_prod):
    composicoes = combo_prod.get('composicoes', [])
    for composicao in composicoes:
        if composicao.get('opcoes'):
            return composicao
    return composicoes[0] if composicoes else {}


def extrair_minimo_do_texto(texto):
    if not texto:
        return None
    match = re.search(r'a\s+partir\s+de\s+(\d+)\s+unidades', texto, re.IGNORECASE)
    if not match:
        match = re.search(r'\((\d+)\s+unidades\)', texto, re.IGNORECASE)
    return int(match.group(1)) if match else None


def numero_ou_none(valor):
    if valor in (None, ""):
        return None
    try:
        return float(str(valor).replace(",", "."))
    except ValueError:
        return None


def obter_preco_opcao(opcao):
    preco_promocional = numero_ou_none(opcao.get('coo_preco_promocional'))
    if preco_promocional is not None:
        return preco_promocional
    return numero_ou_none(opcao.get('coo_preco'))


def obter_quantidade_disponivel_opcao(opcao):
    quantidades = []
    for chave_estoque in ("estoque_opcao", "estoque_opcao_variacao", "estoque_opcao_produto"):
        estoque = opcao.get(chave_estoque)
        if not isinstance(estoque, dict):
            continue

        qtd_estoque = numero_ou_none(estoque.get("iest_qtd"))
        if qtd_estoque is not None:
            quantidades.append(qtd_estoque)

    if not quantidades:
        return None

    return max(0, int(min(quantidades)))


def opcao_disponivel(opcao):
    for chave_estoque in ("estoque_opcao", "estoque_opcao_variacao", "estoque_opcao_produto"):
        estoque = opcao.get(chave_estoque)
        if not isinstance(estoque, dict):
            continue

        if estoque.get("prd_disponivel") == "N":
            return False

        qtd_estoque = estoque.get("iest_qtd")
        if qtd_estoque is None:
            continue

        try:
            if float(qtd_estoque) <= 0:
                return False
        except (TypeError, ValueError):
            continue

    quantidade_disponivel = obter_quantidade_disponivel_opcao(opcao)
    if quantidade_disponivel is not None and quantidade_disponivel <= 0:
        return False

    return True


def filtrar_opcoes_disponiveis(opcoes):
    return [opcao for opcao in opcoes if opcao_disponivel(opcao)]


def obter_preco_a_partir(composicao, min_quantidade):
    opcoes = filtrar_opcoes_disponiveis(composicao.get('opcoes', []))
    precos = [
        preco
        for preco in (obter_preco_opcao(opcao) for opcao in opcoes)
        if preco is not None
    ]
    return round(min(precos) * min_quantidade, 2) if precos else None


def montar_rotulo_combo(nome, min_quantidade):
    nome_limpo = clean_html_entities(nome)
    match_apelido = re.search(r'"([^"]+)"', nome_limpo)
    if match_apelido:
        return f'Combo "{match_apelido.group(1)}" - {min_quantidade} marmitas'

    rotulo = re.sub(r'\s+a\s+partir\s+de\s+\d+\s+unidades.*$', '', nome_limpo, flags=re.IGNORECASE)
    rotulo = re.sub(r'^monte\s+seu\s+combo\s*', 'Combo ', rotulo, flags=re.IGNORECASE).strip()
    if not rotulo.lower().startswith('combo'):
        rotulo = f"Combo {rotulo}"
    return f"{rotulo} - {min_quantidade} marmitas"


def montar_combo_info(combo_prod, fallback=None):
    fallback = fallback or {}
    composicao_principal = obter_composicao_principal(combo_prod)
    produto_id = combo_prod.get("produto_id") or fallback.get("id")
    produto_slug = combo_prod.get("prd_url") or fallback.get("slug")
    nome = combo_prod.get("prd_nome") or fallback.get("nome", "")
    texto_minimo = " ".join([
        str(nome or ""),
        str(composicao_principal.get("com_nome") or ""),
        str(combo_prod.get("prd_descricao") or ""),
    ])

    min_quantidade = int(
        composicao_principal.get("com_qtd_minima")
        or fallback.get("min_quantidade")
        or extrair_minimo_do_texto(texto_minimo)
        or 20
    )
    max_quantidade = int(
        composicao_principal.get("com_qtd")
        or fallback.get("max_quantidade")
        or max(min_quantidade, 1000)
    )
    preco_a_partir = obter_preco_a_partir(composicao_principal, min_quantidade)
    if preco_a_partir is None:
        preco_a_partir = fallback.get("preco_a_partir")

    preco_texto = f"A partir de {formatar_preco_brasileiro(preco_a_partir)}" if preco_a_partir is not None else ""

    return {
        **fallback,
        "id": int(produto_id),
        "nome": nome,
        "rotulo": montar_rotulo_combo(nome, min_quantidade),
        "slug": produto_slug,
        "min_quantidade": min_quantidade,
        "quantidade_padrao": min_quantidade,
        "max_quantidade": max_quantidade,
        "preco_a_partir": preco_a_partir,
        "preco_a_partir_texto": preco_texto,
        "url": montar_combo_url(produto_id, produto_slug),
        "foto": combo_prod.get("prd_foto") or fallback.get("foto"),
    }


def ordenar_combos(combos):
    return sorted(combos, key=lambda combo: (combo.get("min_quantidade", 9999), combo.get("nome", "")))


def obter_url_all_static(html):
    match_script = re.search(r'src=["\'](/cliente/[^"\']/assets/all-static\.js\?v=[0-9]+)["\']', html)
    if not match_script:
        match_script = re.search(r'src=["\'](/cliente/.*?all-static\.js[^"\']*)["\']', html)
    if not match_script:
        raise Exception("Nao foi possivel encontrar a tag do script all-static.js no HTML do site.")
    return BASE_URL + match_script.group(1)


def carregar_produtos_do_site(force=False):
    agora = time.time()
    if (
        not force
        and SITE_DATA_CACHE["produtos"] is not None
        and SITE_DATA_CACHE["expires_at"] > agora
    ):
        return SITE_DATA_CACHE["produtos"]

    url_inicial = FALLBACK_COMBOS[DEFAULT_COMBO_ID]["url"]

    try:
        req = urllib.request.Request(url_inicial, headers=REQUEST_HEADERS)
        with urllib.request.urlopen(req, timeout=10) as response:
            html = response.read().decode('utf-8', errors='ignore')

        script_url = obter_url_all_static(html)
        req_js = urllib.request.Request(script_url, headers=REQUEST_HEADERS)
        with urllib.request.urlopen(req_js, timeout=10) as response:
            js_content = response.read().decode('utf-8', errors='ignore')

        match_produtos = re.search(r'window\.\$_produtos\s*=\s*(.*?);(?=window\.|\Z)', js_content, re.DOTALL)
        if not match_produtos:
            raise Exception("Variavel window.$_produtos nao encontrada no JavaScript do site.")

        produtos = json.loads(match_produtos.group(1).strip())
        SITE_DATA_CACHE["produtos"] = produtos
        SITE_DATA_CACHE["expires_at"] = agora + COMBOS_CACHE_TTL_SECONDS
        return produtos
    except Exception:
        if SITE_DATA_CACHE["produtos"] is not None:
            return SITE_DATA_CACHE["produtos"]
        raise


def descobrir_combos_do_site():
    try:
        produtos = carregar_produtos_do_site()
        combos = []

        for produto in produtos.values():
            nome = produto.get('prd_nome', '')
            if 'monte seu combo' not in nome.lower():
                continue
            if produto.get('prd_exibir_cardapio') == 'N':
                continue

            composicao_principal = obter_composicao_principal(produto)
            if not composicao_principal.get('opcoes'):
                continue

            produto_id = str(produto.get("produto_id"))
            combos.append(montar_combo_info(produto, FALLBACK_COMBOS.get(produto_id)))

        if not combos:
            raise Exception("Nenhum combo foi encontrado no cardapio do site.")

        return ordenar_combos(combos)
    except Exception as exc:
        print(f"Descoberta de combos no site falhou: {exc}. Usando fallback local.")
        return [FALLBACK_COMBOS[combo_id] for combo_id in FALLBACK_COMBO_ORDER]


def combos_publicos():
    return descobrir_combos_do_site()


def get_combo_config(combo_id=None):
    combo_key = str(combo_id or DEFAULT_COMBO_ID)
    for combo in combos_publicos():
        if str(combo["id"]) == combo_key:
            return combo
    raise ValueError("Combo selecionado nao e suportado.")

def extrair_marmitas_do_site(combo_id=None):
    """Faz a raspagem do cardápio direto do site Marmitas da Lulu."""
    try:
        combo_config = get_combo_config(combo_id)
        produtos = carregar_produtos_do_site()
        
        # 5. Encontrar o produto correspondente ao Combo
        combo_prod = produtos.get(str(combo_config["id"]))
        if not combo_prod:
            # Busca resiliente por URL do combo
            for pid, p in produtos.items():
                if p.get('prd_url') == combo_config["slug"]:
                    combo_prod = p
                    break
                    
        if not combo_prod:
            raise Exception("Produto do tipo Combo de Marmitas não encontrado no cardápio do site.")
            
        # 6. Extrair os pratos (opções da composição do combo)
        marmitas = []
        composicoes = combo_prod.get('composicoes', [])
        if not composicoes:
            raise Exception("Composições do combo não encontradas.")
            
        # Geralmente a primeira composição ("Escolha as suas opções...") traz os pratos
        composicao_principal = obter_composicao_principal(combo_prod)
        opcoes = filtrar_opcoes_disponiveis(composicao_principal.get('opcoes', []))
        if not opcoes:
            raise Exception("Nenhuma opção disponível encontrada neste combo.")

        for op in opcoes:
            nome = op.get('coo_nome', '').strip()
            descricao = op.get('coo_complemento', '').strip()
            preco = op.get('coo_preco')
            quantidade_disponivel = obter_quantidade_disponivel_opcao(op)
            
            # Decodificar entidades HTML básicas (ex: &eacute; para é)
            # Como vamos usar a API, podemos repassar para limpar, mas fazemos um replace básico
            descricao = clean_html_entities(descricao)
            
            marmitas.append({
                'id': op.get('composicao_opcao_id'),
                'nome': nome,
                'descricao': descricao,
                'preco': float(preco) if preco is not None else 0.0,
                'foto': op.get('coo_foto'),
                'quantidade_disponivel': quantidade_disponivel,
                'origem': 'site'
            })
            
        combo_info = montar_combo_info(combo_prod, combo_config)

        return marmitas, None, combo_info
    except Exception as e:
        return None, str(e), None

def extrair_marmitas_do_excel():
    """Faz a leitura do cardápio a partir da planilha Excel local."""
    try:
        if not os.path.exists(EXCEL_PATH):
            return None, "Arquivo Excel não encontrado no diretório do projeto."
            
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        if "Opções de marmita" not in wb.sheetnames:
            return None, "Aba 'Opções de marmita' não encontrada na planilha."
            
        sheet = wb["Opções de marmita"]
        marmitas = []
        
        # O cabeçalho é Marmita e Valor na linha 1
        # Linhas de dados começam na linha 2
        for r in range(2, sheet.max_row + 1):
            nome_val = sheet.cell(r, 1).value
            preco_val = sheet.cell(r, 2).value
            
            if nome_val:
                nome = str(nome_val).strip()
                # Limpar o preço
                # Exemplo de preço: "+R$ 28,11" ou "28.11" ou 28.11
                preco = 0.0
                if preco_val is not None:
                    if isinstance(preco_val, (int, float)):
                        preco = float(preco_val)
                    else:
                        # Limpar string
                        preco_str = str(preco_val).replace('+R$', '').replace('R$', '').replace('\xa0', '').strip()
                        preco_str = preco_str.replace('.', '').replace(',', '.') # Converter padrão PT-BR para float
                        try:
                            preco = float(preco_str)
                        except ValueError:
                            preco = 0.0
                            
                marmitas.append({
                    'id': f"xls_{r}",
                    'nome': nome,
                    'descricao': "Disponível na planilha local",
                    'preco': preco,
                    'foto': None,
                    'quantidade_disponivel': None,
                    'origem': 'excel'
                })
                
        return marmitas, None
    except Exception as e:
        return None, str(e)

def clean_html_entities(text):
    """Substitui algumas entidades HTML básicas comuns no site."""
    if not text:
        return ""
    replacements = {
        "&ldquo;": '"', "&rdquo;": '"', "&lsquo;": "'", "&rsquo;": "'",
        "&nbsp;": " ", "&amp;": "&", "&lt;": "<", "&gt;": ">",
        "&quot;": '"', "&#039;": "'", "&aacute;": "á", "&eacute;": "é",
        "&iacute;": "í", "&oacute;": "ó", "&uacute;": "ú",
        "&atilde;": "ã", "&otilde;": "õ", "&acirc;": "â", "&ecirc;": "ê",
        "&ocirc;": "ô", "&ccedil;": "ç", "&Aacute;": "Á", "&Eacute;": "É",
        "&Iacute;": "Í", "&Oacute;": "Ó", "&Uacute;": "Ú",
        "&Atilde;": "Ã", "&Otilde;": "Õ", "&Ccedil;": "Ç"
    }
    for ent, char in replacements.items():
        text = text.replace(ent, char)
    # Remover tags HTML remanescentes se houver
    text = re.sub(r'<[^>]+>', '', text)
    return text


def normalizar_texto_busca(texto):
    return clean_html_entities(str(texto or "")).strip().lower()


def obter_limite_item(marmita, max_repeticoes):
    quantidade_disponivel = marmita.get("quantidade_disponivel")
    if quantidade_disponivel in (None, ""):
        return max_repeticoes

    try:
        return min(max_repeticoes, max(0, int(float(quantidade_disponivel))))
    except (TypeError, ValueError):
        return max_repeticoes


def validar_sugestao_gerada(resposta_json, marmitas, quantidade, max_repeticoes):
    itens = resposta_json.get("itens_selecionados", [])
    marmitas_por_id = {str(m.get("id")): m for m in marmitas}
    marmitas_por_nome = {normalizar_texto_busca(m.get("nome")): m for m in marmitas}
    erros = []
    total = 0

    for item in itens:
        item_id = item.get("id") or item.get("opcao_id") or item.get("composicao_opcao_id")
        marmita = marmitas_por_id.get(str(item_id)) if item_id not in (None, "") else None
        if marmita is None:
            marmita = marmitas_por_nome.get(normalizar_texto_busca(item.get("nome")))

        try:
            qtd = int(item.get("quantidade") or 0)
        except (TypeError, ValueError):
            qtd = 0

        total += qtd

        if marmita is None:
            erros.append(f"{item.get('nome', 'Item sem nome')} nao foi encontrado no cardapio atual.")
            continue

        limite = obter_limite_item(marmita, max_repeticoes)
        if qtd > limite:
            estoque = marmita.get("quantidade_disponivel")
            erros.append(
                f"{marmita.get('nome')} foi escolhido {qtd}x, mas o limite atual e {limite} "
                f"(estoque: {estoque if estoque is not None else 'nao informado'})."
            )

        item["id"] = marmita.get("id")

    if total != quantidade:
        erros.append(f"A soma da sugestao ficou em {total} marmitas, mas o pedido exige {quantidade}.")

    return erros

@app.route('/')
def index():
    return render_template('index.html', combos=combos_publicos(), default_combo_id=DEFAULT_COMBO_ID)

@app.route('/instalar-automacao-carrinho')
def instalar_automacao_carrinho():
    script_path = os.path.join(app.static_folder, 'js', 'lulu-cart.user.js')
    with open(script_path, 'r', encoding='utf-8') as script_file:
        script_content = script_file.read()
    return render_template('install_userscript.html', script_content=script_content)

@app.route('/api/cardapio')
def api_cardapio():
    combo_id = request.args.get('combo_id', DEFAULT_COMBO_ID)
    try:
        combo_config = get_combo_config(combo_id)
    except ValueError as error:
        return jsonify({
            'sucesso': False,
            'erro': str(error)
        }), 400

    # Tenta raspar do site primeiro
    marmitas, erro_site, combo_info = extrair_marmitas_do_site(combo_id)
    
    if marmitas:
        return jsonify({
            'sucesso': True,
            'origem': 'site',
            'combo': combo_info,
            'marmitas': marmitas
        })
        
    # Se falhar, tenta ler do Excel como fallback
    print(f"Raspagem do site falhou: {erro_site}. Tentando ler do Excel...")
    if str(combo_config["id"]) != DEFAULT_COMBO_ID:
        return jsonify({
            'sucesso': False,
            'erro': f"Erro ao acessar site ({erro_site}). A planilha local disponivel e apenas do combo de 20 unidades."
        }), 500

    marmitas_xls, erro_xls = extrair_marmitas_do_excel()
    
    if marmitas_xls:
        return jsonify({
            'sucesso': True,
            'origem': 'excel',
            'combo': combo_config,
            'alerta': f"Erro ao acessar site ({erro_site}). Dados carregados do Excel local.",
            'marmitas': marmitas_xls
        })
        
    return jsonify({
        'sucesso': False,
        'erro': f"Não foi possível carregar os dados. Site: {erro_site} | Excel: {erro_xls}"
    }), 500

@app.route('/api/gerar-sugestao', methods=['POST'])
def api_gerar_sugestao():
    dados = request.json
    api_key = dados.get('api_key')
    marmitas = dados.get('marmitas', [])
    criterios = dados.get('criterios', {})
    
    if not api_key:
        return jsonify({'sucesso': False, 'erro': 'A chave de API do Gemini é obrigatória.'}), 400
        
    if not marmitas:
        return jsonify({'sucesso': False, 'erro': 'Nenhuma marmita informada para seleção.'}), 400
        
    try:
        # Configurar API do Gemini
        genai.configure(api_key=api_key)
        
        # Parâmetros de critérios
        objetivo = criterios.get('objetivo', 'Saudável / Variado')
        quantidade = int(criterios.get('quantidade', 20))
        max_repeticoes = int(criterios.get('max_repeticoes', 3))
        valor_maximo = criterios.get('valor_maximo')
        exclusoes = criterios.get('exclusoes', '')
        preferencias = criterios.get('preferencias', '')
        combo_info = criterios.get('combo') or {}
        combo_nome = combo_info.get('nome') or FALLBACK_COMBOS[DEFAULT_COMBO_ID]["nome"]
        combo_minimo = int(combo_info.get('min_quantidade') or FALLBACK_COMBOS[DEFAULT_COMBO_ID]["min_quantidade"])

        if quantidade < combo_minimo:
            return jsonify({
                'sucesso': False,
                'erro': f'O combo selecionado exige pelo menos {combo_minimo} marmitas.'
            }), 400

        capacidade_total = sum(obter_limite_item(marmita, max_repeticoes) for marmita in marmitas)
        if capacidade_total < quantidade:
            return jsonify({
                'sucesso': False,
                'erro': f'Com o estoque atual e o limite de repeticoes, so e possivel montar {capacidade_total} marmitas.'
            }), 400
        
        # Obter a saudação dinâmica baseada no horário local do servidor
        from datetime import datetime
        hora_atual = datetime.now().hour
        if 5 <= hora_atual < 12:
            saudacao = "Bom dia!"
        elif 12 <= hora_atual < 18:
            saudacao = "Boa tarde!"
        else:
            saudacao = "Boa noite!"

        # Formatar a lista de marmitas para o prompt
        lista_marmitas_str = ""
        for idx, m in enumerate(marmitas):
            estoque = m.get('quantidade_disponivel')
            estoque_texto = f"{estoque} unidades" if estoque is not None else "nao informado"
            limite_efetivo = obter_limite_item(m, max_repeticoes)
            lista_marmitas_str += (
                f"[{idx}] ID_OPCAO: {m['id']} | {m['nome']} - Preço: R$ {float(m['preco']):.2f} "
                f"| Estoque disponivel: {estoque_texto} | Maximo que pode escolher deste prato: {limite_efetivo} "
                f"| Ingredientes/Descrição: {m['descricao']}\n"
            )
            
        prompt = f"""
Você é um nutricionista especialista e assistente de delivery. Seu objetivo é ajudar o usuário a escolher a combinação ideal de marmitas saudáveis a partir do cardápio do site "Marmitas da Lulu".

Aqui está a lista de marmitas disponíveis no cardápio atualmente:
{lista_marmitas_str}

CRITÉRIOS DO USUÁRIO:
- Combo selecionado: {combo_nome}
- Objetivo Nutricional: {objetivo}
- Quantidade total de marmitas a selecionar: {quantidade} unidades.
- Quantidade máxima de repetição de cada prato: {max_repeticoes} unidades (não pedir mais do que {max_repeticoes} do mesmo tipo).
{f"- Valor total máximo aproximado: R$ {valor_maximo:.2f}" if valor_maximo else "- Orçamento: Sem limite específico (escolher as melhores de acordo com a nutrição)"}
- Ingredientes a serem EXCLUÍDOS ou restrições alimentares: {exclusoes if exclusoes else "Nenhuma restrição especial"}
- Preferências alimentares específicas: {preferencias if preferencias else "Nenhuma preferência especial"}

SUAS DIRETRIZES DE SELEÇÃO:
1. QUANTIDADE EXATA: A soma das quantidades de todas as marmitas escolhidas deve ser EXATAMENTE igual a {quantidade}. Nem mais, nem menos.
2. MÁXIMO DE REPETIÇÕES E ESTOQUE: Você NÃO pode selecionar mais do que {max_repeticoes} unidades de uma única marmita e também NÃO pode ultrapassar o estoque disponível de cada prato. O máximo efetivo de cada prato é o menor valor entre o limite de repetição e o estoque disponível informado no cardápio.
3. ORÇAMENTO: Se houver um valor total máximo aproximado, a soma dos preços de todas as marmitas escolhidas deve ficar dentro ou o mais próximo possível desse valor. Lembre-se de multiplicar a quantidade de cada prato pelo seu preço individual na soma.
4. OBJETIVO NUTRICIONAL: Escolha os pratos adequados ao objetivo. Se for "Low Carb", prefira marmitas da linha Low Carb ou com poucos carboidratos. Se for "Ganho de Massa / Maromba", prefira opções proteicas e com carboidratos complexos (linha Maromba). Se for "Emagrecimento", prefira pratos menos calóricos.
5. EXCLUSÕES: Não selecione NENHUMA marmita que contenha ingredientes listados nas restrições/exclusões do usuário. Analise bem o nome e a descrição do prato para garantir isso.
6. PREFERÊNCIAS: Dê peso maior às marmitas que se alinhem às preferências do usuário.

FORMATO DE RETORNO ESPERADO (Gere APENAS um JSON no formato especificado abaixo, sem crases de formatação markdown e sem textos adicionais antes ou depois. O JSON deve ser diretamente interpretável por `json.loads`):
{{
  "itens_selecionados": [
    {{
      "id": "ID_OPCAO_EXATO_DO_CARDAPIO",
      "nome": "NOME DO PRATO EXATO",
      "quantidade": 3,
      "preco_unitario": 28.11,
      "subtotal": 84.33
    }}
  ],
  "valor_total": 560.20,
  "justificativa_nutricional": "Uma breve explicação em português do Brasil (1 parágrafo) de como essa seleção atende ao objetivo nutricional, limite de repetições e preferências do usuário.",
  "mensagem_whatsapp": "A mensagem formatada pronta para enviar no WhatsApp para fazer o pedido. Deve ser cortês, clara e estruturada.\\n\\nREGRAS CRÍTICAS DE FORMATAÇÃO DA MENSAGEM DO WHATSAPP:\\n1. Comece a mensagem exatamente com: '{saudacao} Tudo bem?' e uma quebra de linha.\\n2. NUNCA cite a palavra 'Lulu' ou use nomes próprios de atendentes.\\n3. NÃO coloque o valor total calculado na mensagem do WhatsApp, pois o usuário deseja que o atendimento do delivery informe o valor oficial. Remova qualquer menção ao preço total ou subtotais.\\n\\nExemplo de estrutura da mensagem:\\n{saudacao} Tudo bem?\\n\\nGostaria de fazer o pedido do meu combo de {quantidade} marmitas:\\n\\n- 3x LINHA MAROMBA: SOBRECOXA...\\n- 2x LINHA LOW CARB: PATINHO...\\n\\nMuito obrigado!\\n\\n(A mensagem deve listar apenas os pratos e quantidades exatas no formato acima, sem preços e sem totais)"
}}
"""

        # Chamar o modelo Gemini 2.5 Flash
        # O modelo recomendado para tarefas rápidas e estruturadas é 'gemini-2.5-flash'
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        response = model.generate_content(
            prompt,
            generation_config=genai.types.GenerationConfig(
                response_mime_type="application/json"
            )
        )
        
        # Fazer o parse da resposta
        resposta_json = json.loads(response.text.strip())
        erros_sugestao = validar_sugestao_gerada(resposta_json, marmitas, quantidade, max_repeticoes)
        if erros_sugestao:
            return jsonify({
                'sucesso': False,
                'erro': "O Gemini gerou uma seleção fora dos limites atuais de estoque. Tente gerar novamente.",
                'detalhe': " | ".join(erros_sugestao)
            }), 500

        return jsonify({
            'sucesso': True,
            'dados': resposta_json
        })
        
    except json.JSONDecodeError as jde:
        print("Erro ao decodificar JSON retornado pelo Gemini:", response.text)
        return jsonify({
            'sucesso': False,
            'erro': "O Gemini gerou uma resposta, mas ela não veio no formato JSON correto. Tente gerar novamente.",
            'detalhe': str(jde),
            'resposta_crua': response.text if 'response' in locals() else None
        }), 500
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': f"Erro ao processar requisição com o Gemini: {str(e)}"
        }), 500

if __name__ == '__main__':
    # Criar diretórios se não existirem
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static/css', exist_ok=True)
    os.makedirs('static/js', exist_ok=True)
    
    print("Iniciando servidor local do QueroPedirMarmitas...")
    app.run(debug=True, port=5000)
